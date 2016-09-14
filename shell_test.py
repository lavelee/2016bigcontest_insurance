# -*- coding: utf-8 -*-

#리눅스에서 딥러닝 모듈을 돌리기 위한 쉘파일
import subprocess
import os
import pickle
import time
import openpyxl
import numpy

t1=time.localtime()
print 'start at', time.strftime('%y%m%d %Hh%Mm',t1)

#pickle_file = sys.argv[1] #full picklefile path
#layer2_nodes =  sys.argv[2] #100 or 1024
#learning_rate_init= sys.argv[3] #0.5

#이 shell은 타겟파일에서 순서대로 피클네임, 노드수, 학습율 을 변수로 받고 결과를 pcrf 순으로 내보내는데 맞춰서 작성되었다. 

folder = '/notebooks/insurance/pickle_files/' #실제 실행용 pickle folder
subpy = '/notebooks/insurance/Insurance_model.py' #쉘로 실행할 py 파일의 위치
#folder = '/home/rbl/Documents/TensorFlow/insurance/pickle_files_test/' #디버깅용으로 파일 2개만 넣어서 돌려봄
#변인수를 변수로 두는것은 총 개수를 동적으로 계산해 몇개중 몇번째것이 진행중인지 표시하기 위해서임.
n_node_var=4 # 노드 변인수. 변인을 이 값에 연동시켰으므로 변인수 바뀌면 다시 만들어야 함. 0부터 세지 않음.
n_Lrate_var=2 # 러닝레이트 변인수
n_test_per_var=5 # 한 조건세트당 몇번씩 수행
select = 'cuclaim' #cuclaim, cucntt 중에 고를수 있게



n_total_test = len(os.listdir(folder))*n_node_var*n_Lrate_var*n_test_per_var #총 횟수
test_result = [['picklename','select','nodes','L rate','predict','correct predict','real','F1 score']] #최종출력 첫행에 컬럼명 넣기 


def pcrfCollector(book):
#sin predict, sin correct predict, sin real, sin f1 score 부분을 텍스트에서 찾고 오른쪽의 값을 리턴한다
    targets =['\nsin predict : ','\nsin correct predict : ','\nsin real : ','\nsin F1 score : ','\n\npicklize finished']
    result =[] #위 targets 배열의 마지막은 stopper 로 실제 탐색에 사용되진 않지만 멈추는 위치 지정이다. 
    for i in range(0,len(targets)-1) : #마지막 배열은 ending point 라서 제거 
        start = book.find(targets[i])+len(targets[i])
        end = book.find(targets[i+1])
        result.append(float(book[start:end]))
    #print result
    return result

def where(h,i,j,k,imax=n_node_var,jmax=n_Lrate_var,kmax=n_test_per_var): 
#현재위치 셈. hmax는 필요가 없다
    position = (h*n_node_var*n_Lrate_var*n_test_per_var +
                    i*                 n_Lrate_var*n_test_per_var +
                    j*                                   n_test_per_var +
                    k+1                                                       ) #마지막 digit에는 +1 해줘야함.
    return position


def pickleread(pickle_name):
    with open(pickle_name,'rb') as f:
        data=pickle.load(f)
        return data

def sheetmake(data,excel_name):
    book = openpyxl.Workbook()
    for dictitle , dictdata in data.items():
        dictdata=numpy.matrix(dictdata) #1차원 배열 있으면 shape 차원 하나라 오류나서.
        sheet=book.create_sheet(title=dictitle)
        for n_col in range(0,dictdata.shape[1]):
            for n_row in range(0,dictdata.shape[0]):
                #input_value=numpy.asscalar(dictdata[n_row,n_col])          #python native 로 바꿔주는 코드. 이것과 아래줄 둘중하나 필수. 
                input_value=dictdata[n_row,n_col]                           #str 오류날때 asscalar 빼면 될때있음
                sheet.cell(row=n_row+1,column=n_col+1).value=input_value    #엑셀에선 행,열 첫번호가 1 
            #sheet.column_dimensions[openpyxl.cell.get_column_letter(n_col+1)].width = 2.76 #컬럼 넓이 조절. 필요 없으면 빼기
        print 'making sheet : ',dictitle
    sheet = book.get_sheet_by_name('Sheet') #select sheet named Sheet
    book.remove_sheet(sheet) #delete that sheet
    print 'saving data to excel...'
    book.save(excel_name)
    print 'finished, file saved : ',excel_name


#시작
for h, filename in enumerate(os.listdir(folder)):
    pickle_file = os.path.join(folder,filename)
    for i in range(0,n_node_var):
        if i==0:
            layer2_nodes = 10
        elif i==1:
            layer2_nodes = 50
        elif i==2:
            layer2_nodes = 100
        elif i==3:
            layer2_nodes = 1024
        for j in range(0,n_Lrate_var): #현재 learning_rate 는 0.5로 고정이니 1개라서 range(0,1) 
            if j==0:
                learning_rate_init = 0.3
            if j==1:
                learning_rate_init = 0.1
            for k in range(0,n_test_per_var): #같은 피클과 조건에 대해 몇번 반복할것인가 
                syscommand = 'python '+subpy+' "'+pickle_file+'" '+str(layer2_nodes)+' '+str(learning_rate_init)+' '+str(k+1)+' '+select
                #print '\n',syscommand
                get=subprocess.check_output(syscommand, shell=True)
                output = pcrfCollector(get)
                output.insert(0,learning_rate_init) #리턴값이 insert된 배열이 아니라.. 이거 실행만으로 insert 되는 함수임. 
                output.insert(0,layer2_nodes)
                output.insert(0,select)
                output.insert(0,filename) #마지막에 0에 넣는게 가장 앞으로 오니까. 
                #print(h,i,j,k)
                print where(h,i,j,k), '/' ,n_total_test, output #현재 위치와 최대값 표시
                test_result.append(output) #이것도 재할당 안해도 자동 적용되는거. 
#output [피클명 , nodes , learningrate, predict, correct predict, real, f1score]
#print '\n', test_result    #최종 출력배열 확인



#피클화. 날짜 시간 folderprocess.pickle 로 저장
pickle_name = time.strftime('%y%m%d %Hh%Mm',t1) + ' FolderResult.pickle'
f = open(pickle_name,'w')
save={
    'result' : test_result        
        }
pickle.dump(save,f,pickle.HIGHEST_PROTOCOL)
f.close()
#print '\npicklize finished.  filename :',pickle_name


#엑셀파일 저장
excel_name=folder+pickle_name[:pickle_name.find(".pickle")]+'.xlsx' #folder 붙여서 피클파일있던 폴더안에 저장하게
if os.path.isfile(excel_name): #이미 파일이 있으면 삭제함 #엑셀파일이 열려있으면 삭제도 못하고 오류남. 
    os.remove(excel_name)
    print('target excel file exists. continue after deleting')
sheetmake(pickleread(pickle_name),excel_name) #피클읽어 그대로 sheet 로 출력. 
os.remove(pickle_name)#피클 제거


t2=time.localtime() #끝난 시간
print 'start at', time.strftime('%y%m%d %Hh%Mm',t1)
print 'ends at',  time.strftime('%y%m%d %Hh%Mm',t2)


