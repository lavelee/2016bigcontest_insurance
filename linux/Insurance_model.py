# coding: utf-8
#리눅스에서 쉘파일로 실행되는 머신러닝 코드. 윈도우에서는 실행되지 않는다. 자체파일로만은 받아야할 변수가 없어 실행 안된다. 

from __future__ import print_function
import numpy as np
import tensorflow as tf
from six.moves import cPickle as pickle
import sys, os
import openpyxl

pickle_file = sys.argv[1] #full picklefile path
layer2_nodes =  int(sys.argv[2]) #100 or 1024
learning_rate= float(sys.argv[3]) #0.5
tryno = int(sys.argv[4])
num_labels = 2

batch_size = 100
w_init_deviation = 0.5
keep_ratio=1


path = pickle_file[:pickle_file.rfind('/')+1] #피클파일 저장된 경로 . 생성 결과파일 저장할때 경로로 사용
filename = pickle_file[pickle_file.rfind('/')+1:pickle_file.rfind('.pickle')] #전체경로에서 피클파일 이름만. 확장자 제거하고 
resultname = path+'train) '+filename+' node'+str(layer2_nodes)+' LR'+str(learning_rate)+' try'+str(tryno)#폴더경로에 파일저장. 
#print(resultname)
pickle_name=resultname+'.pickle'
excel_name=resultname+'.xlsx' #확장자만 바꿔줌
#피클과 엑셀이름  :  train) 피클이름 노드수 러닝레이트 트레인횟수 
# train) rand1 node100 LR0.5 try1 .xlsx
#한번 시행당 파일 3개가 만들어진다. 하나는 텐서플로 variable 방식 w,b저장 확장자없는것,  pickle로 w,b 저장, 엑셀로 w,b 출력한 3가지다. 


with open(pickle_file, 'rb') as f:
  save = pickle.load(f)
  test_labels = save['test_label']
  test_dataset = save['test_data']
  train_labels = save['train_label']
  train_dataset = save['train_data']
  column_names = save['col_names']
  train_distri = save['train_distri']
  submit_dataset = save['submit_data']
  submit_custid = save['submit_custid']
del save  #to free up memory


def reformat(labels):
  one_hot_encode = (np.arange(num_labels) == labels[:,None]).astype(np.float32)
  return one_hot_encode 

test_labels=reformat(test_labels)
train_labels=reformat(train_labels)



def accuracy(predictions, labels):
  return (100.0 * np.sum(np.argmax(predictions, 1) == np.argmax(labels, 1))
          / predictions.shape[0])

def f1Score(predictions, labels):
    #print ('predictions_raw=',predictions)
    predictions = np.array(predictions==predictions.max(axis=1)[:,None],dtype='int')
    predict = np.sum(predictions,0)
    real = np.sum(labels,0)
    correct_predict = np.sum(labels*(np.argmax(predictions, 1) == np.argmax(labels, 1))[:,None],0)
    precision = correct_predict[1]/predict[1]
    recall = correct_predict[1]/real[1]
    f1_score=2*precision*recall/(precision+recall)
    print('sin predict :',predict[1])
    print('sin correct predict :',correct_predict[1])
    print('sin real :',real[1])
    print('sin F1 score :',f1_score)


step_size= train_dataset.shape[0]//batch_size
graph = tf.Graph()
with graph.as_default():
  tf_train_dataset = tf.placeholder(tf.float32,shape=(batch_size, train_dataset.shape[1]))
  tf_train_labels = tf.placeholder(tf.float32, shape=(batch_size, num_labels))
  tf_test_dataset = tf.constant(test_dataset)
    
  L1_weights = tf.Variable(
    tf.truncated_normal([int(train_dataset.shape[1]), layer2_nodes], mean=0.0, stddev=w_init_deviation))
  L1_biases =  tf.Variable(tf.zeros([layer2_nodes]))
  L1_logits= tf.nn.dropout(tf.matmul(tf_train_dataset, L1_weights) + L1_biases,keep_ratio)
    
  L2_weights = tf.Variable(
    tf.truncated_normal([layer2_nodes, num_labels], mean=0.0, stddev=w_init_deviation)) #다음레이어는 classifier 니까 class개수.
  L2_biases = tf.Variable(tf.zeros([num_labels]))
  L2_logits=tf.matmul(tf.nn.relu(L1_logits), L2_weights) + L2_biases #마지막 레이어는 dropout 하지않음
    
  loss = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(L2_logits, tf_train_labels))
    
  optimizer = tf.train.GradientDescentOptimizer(learning_rate).minimize(loss)
    
  train_prediction = tf.nn.softmax(L2_logits)
  test_prediction =tf.nn.softmax(tf.matmul(tf.nn.relu(tf.matmul(test_dataset, L1_weights) + L1_biases),L2_weights)+L2_biases)
  submit_prediction =tf.nn.softmax(tf.matmul(tf.nn.relu(tf.matmul(submit_dataset, L1_weights) + L1_biases),L2_weights)+L2_biases)
    
  saver = tf.train.Saver({'w1' : L1_weights,
                          'b1' : L1_biases,
                          'w2' : L2_weights,
                          'b2' : L2_biases,
                         })


with tf.Session(graph=graph) as session:
  tf.initialize_all_variables().run()
  print("Initialized")
  for step in range(step_size):
    offset = (step * batch_size) % (train_labels.shape[0] - batch_size)
    batch_data = train_dataset[offset:(offset + batch_size), :]
    batch_labels = train_labels[offset:(offset + batch_size), :]
    feed_dict = {tf_train_dataset : batch_data, tf_train_labels : batch_labels}
    global_step=step
    _,l, predictions = session.run(
      [optimizer,loss, train_prediction], feed_dict=feed_dict)
  f1Score(test_prediction.eval(), test_labels)
  saver.save(session,resultname+'_tensorflow',write_meta_graph=False) 
  w1 = session.run(L1_weights)
  b1 = session.run(L1_biases)
  w2 = session.run(L2_weights)
  b2 = session.run(L2_biases)
  submit_result = np.concatenate((submit_custid,submit_prediction.eval()),1)
  print(submit_result.shape)
  print(submit_result)


f=open(pickle_name,'w')
save ={'w1' : w1,
       'b1' : b1,
       'w2' : w2,
       'b2' : b2,
       'col_names' : np.array(column_names),#numpy.array 안붙이면 shape 못쓰기도 하고 다음에 꺼내쓸때 오류날수있음
       'tr_dist' : np.array(train_distri),
       'submit' : np.array(submit_result)
      }
pickle.dump(save,f,pickle.HIGHEST_PROTOCOL)
f.close()

def pickleread(pickle_name):
    with open(pickle_name,'rb') as f:
        data=pickle.load(f)
        return data

def sheetmake(data,excel_name):
    book = openpyxl.Workbook()
    for dictitle , dictdata in data.items():
        dictdata=np.matrix(dictdata) #1차원 배열 있으면 shape 차원 하나라 오류나서.
        sheet=book.create_sheet(title=dictitle)
        for n_col in range(0,dictdata.shape[1]):
            for n_row in range(0,dictdata.shape[0]):
                input_value=np.asscalar(dictdata[n_row,n_col])          #python native 로 바꿔주는 코드. 이것과 아래줄 둘중하나 필수. 
                #input_value=dictdata[n_row,n_col]                           #str 오류날때 asscalar 빼면 될때있음
                sheet.cell(row=n_row+1,column=n_col+1).value=input_value    #엑셀에선 행,열 첫번호가 1 
            sheet.column_dimensions[openpyxl.cell.get_column_letter(n_col+1)].width = 2.76 #컬럼 넓이 조절. 필요 없으면 빼기
        print('making sheet : ',dictitle)
    sheet = book.get_sheet_by_name('Sheet') #select sheet named Sheet
    book.remove_sheet(sheet) #delete that sheet
    print('saving data to excel...')
    book.save(excel_name)
    print ('finished, file saved : ',excel_name)

excel_name=resultname+'.xlsx' 
sheetmake(pickleread(pickle_name),excel_name)


