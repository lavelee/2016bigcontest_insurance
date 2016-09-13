# coding: utf-8
from __future__ import print_function
import numpy as np
import tensorflow as tf
from six.moves import cPickle as pickle
import sys
import openpyxl

pickle_file = sys.argv[1] #full picklefile path
layer2_nodes =  int(sys.argv[2]) #100 or 1024
learning_rate_init= float(sys.argv[3]) #0.5
tryno = int(sys.argv[4])
select = 'cuclaim'#select cucntt or cuclaim

path = pickle_file[:pickle_file.rfind('/')+1]
filename = pickle_file[pickle_file.rfind('/')+1:pickle_file.rfind('.pickle')]
resultname = path+'train) '+filename+' node'+str(layer2_nodes)+' LR'+str(learning_rate_init)+' try'+str(tryno)+' '+select #폴더경로에 파일저장. 

print (pickle_file)
print (layer2_nodes)
print (learning_rate_init)

#pickle_file = '/home/rbl/Documents/TensorFlow/insurance/pickle_files/cucntt_cuclaim_null_randomfix.pickle'
#pickle_file = '/pickle_files/cucntt_cuclaim_null_randomfix.pickle' # doesn't work
#pickle_file = 'cucntt_cuclaim_null_randomfix.pickle' #same directory

with open(pickle_file, 'rb') as f:
  save = pickle.load(f)
  test_cucntt_label = save['test_cucntt_label']
  test_cucntt_data = save['test_cucntt_data']
  train_cucntt_label = save['train_cucntt_label']
  train_cucntt_data = save['train_cucntt_data']
  test_cuclaim_label = save['test_cuclaim_label']
  test_cuclaim_data = save['test_cuclaim_data']
  train_cuclaim_label = save['train_cuclaim_label']
  train_cuclaim_data = save['train_cuclaim_data']
  cucntt_column_names = save['cucntt_column_names']
  cuclaim_column_names = save['cuclaim_column_names']
  train_cucntt_distri = save['train_cucntt_distri']
  train_cuclaim_distri = save['train_cuclaim_distri']
  

del save  # hint to help gc free up memory
#print('cucntt Test set : ', test_cucntt_label.shape, test_cucntt_data.shape)
#print('cucntt Training set : ', train_cucntt_label.shape, train_cucntt_data.shape)
#print('cuclaim Test set : ', test_cuclaim_label.shape, test_cuclaim_data.shape)
#print('cuclaim Training set : ', train_cuclaim_label.shape, train_cuclaim_data.shape)


# one hot encoding
num_labels = 2
def reformat(labels):
  one_hot_encode = (np.arange(num_labels) == labels[:,None]).astype(np.float32)
  return one_hot_encode 

test_cucntt_label=reformat(test_cucntt_label)
train_cucntt_label=reformat(train_cucntt_label)
test_cuclaim_label=reformat(test_cuclaim_label)
train_cuclaim_label=reformat(train_cuclaim_label)

#print('cucntt Test set : ', test_cucntt_label.shape, test_cucntt_data.shape)
#print('cucntt Training set : ', train_cucntt_label.shape, train_cucntt_data.shape)
#print('cuclaim Test set : ', test_cuclaim_label.shape, test_cuclaim_data.shape)
#print('cuclaim Training set : ', train_cuclaim_label.shape, train_cuclaim_data.shape)


# In[4]:

def accuracy(predictions, labels):
  return (100.0 * np.sum(np.argmax(predictions, 1) == np.argmax(labels, 1))
          / predictions.shape[0])

def f1Score(predictions, labels):
    predictions = np.array(predictions==predictions.max(axis=1)[:,None],dtype='int')
    predict = np.sum(predictions,0)
    real = np.sum(labels,0)
    correct_predict = np.sum(labels*(np.argmax(predictions, 1) == np.argmax(labels, 1))[:,None],0)
    f1_score=2*correct_predict[1]/predict[1]*correct_predict[1]/real[1]/(correct_predict[1]/predict[1]+correct_predict[1]/real[1])
    #print(predictions)
    #print(labels)
    #print(predict)
    print("sin predict :",predict[1])
    print("sin correct predict :",correct_predict[1])
    print("sin real :",real[1])
    #print("\nnot_sin pricision : ",correct_predict[0]/predict[0])
    #print("not_sin   recall  : ",correct_predict[0]/real[0])
    #print("not_sin F1 score  : ",2*correct_predict[0]/predict[0]*correct_predict[0]/real[0]/(correct_predict[0]/predict[0]+correct_predict[0]/real[0]))
    #print("sin pricision : ",correct_predict[1]/predict[1])
    #print("sin   recall :",correct_predict[1]/real[1])
    print("sin F1 score :",f1_score)


if select=="cucntt":
    train_dataset = train_cucntt_data
    train_labels = train_cucntt_label
    test_dataset = test_cucntt_data
    test_labels = test_cucntt_label
elif select=="cuclaim":
    train_dataset = test_cuclaim_data
    train_labels = train_cuclaim_label
    test_dataset = test_cuclaim_data
    test_labels = test_cuclaim_label
else : 
    raise error
#print (select,' selected')


# In[5]:

#setting py
batch_size = 100
beta = 0
w_init_deviation = 0.5
#learning_rate_init=0.5
l_rate_final_ratio=1
#layer2_nodes = 100
keep_ratio=1


step_size= train_dataset.shape[0]//batch_size
#print('step_size = ',step_size)

graph = tf.Graph()
with graph.as_default():
  tf_train_dataset = tf.placeholder(tf.float32,shape=(batch_size, train_dataset.shape[1]))
  tf_train_labels = tf.placeholder(tf.float32, shape=(batch_size, num_labels))
  tf_test_dataset = tf.constant(test_dataset)
    
  L1_weights = tf.Variable(
    tf.truncated_normal([int(train_dataset.shape[1]), layer2_nodes], mean=0.0, stddev=w_init_deviation))
  L1_biases =  tf.Variable(tf.zeros([layer2_nodes]))
  L1_logits= tf.nn.dropout(tf.matmul(tf_train_dataset, L1_weights) + L1_biases,keep_ratio)
    
  L2_weights = tf.Variable(
    tf.truncated_normal([layer2_nodes, num_labels], mean=0.0, stddev=w_init_deviation))
  L2_biases = tf.Variable(tf.zeros([num_labels]))
  L2_logits=tf.nn.dropout(tf.matmul(tf.nn.relu(L1_logits), L2_weights) + L2_biases,keep_ratio)
    
  loss = tf.reduce_mean(
    tf.nn.softmax_cross_entropy_with_logits(L2_logits, tf_train_labels) + 
    beta*(tf.nn.l2_loss(L1_weights) + tf.nn.l2_loss(L2_weights)))
    
  global_step = tf.Variable(0, trainable=False)  
  learning_rate=tf.train.exponential_decay(learning_rate_init, global_step, step_size, l_rate_final_ratio, staircase=False)
  optimizer = tf.train.GradientDescentOptimizer(learning_rate).minimize(loss,global_step=global_step)
    
  train_prediction = tf.nn.softmax(L2_logits)
  test_prediction =tf.nn.softmax(
    tf.matmul(tf.nn.relu(tf.matmul(test_dataset, L1_weights) + L1_biases),L2_weights)+L2_biases)
    
  saver = tf.train.Saver({'w1' : L1_weights,
                          'b1' : L1_biases,
                          'w2' : L2_weights,
                          'b2' : L2_biases,
                         })


# In[6]:

with tf.Session(graph=graph) as session:
  tf.initialize_all_variables().run()
  #print("Initialized")
  for step in range(step_size):
    offset = (step * batch_size) % (train_labels.shape[0] - batch_size)
    batch_data = train_dataset[offset:(offset + batch_size), :]
    batch_labels = train_labels[offset:(offset + batch_size), :]
    feed_dict = {tf_train_dataset : batch_data, tf_train_labels : batch_labels}
    global_step=step
    _,l, predictions = session.run(
      [optimizer,loss, train_prediction], feed_dict=feed_dict)
    #if (step % (int(step_size*0.2)) == 0):
      #print("Minibatch loss at step %d: %f" % (step, l))
      #print("Minibatch accuracy: %.1f%%" % accuracy(predictions, batch_labels))
  #print("Test accuracy: %.1f%%" % accuracy(test_prediction.eval(), test_labels))
  #print("\nby ",select)
  f1Score(test_prediction.eval(), test_labels)
  saver.save(session,resultname,write_meta_graph=False)
  w1 = session.run(L1_weights)
  b1 = session.run(L1_biases)
  w2 = session.run(L2_weights)
  b2 = session.run(L2_biases)


pickle_name=resultname+'.pickle'
f=open(pickle_name,'w')
save ={'w1' : w1,
       'b1' : b1,
       'w2' : w2,
       'b2' : b2,
       'cucntt_cn' : np.array(cucntt_column_names),   #numpy.array 안붙이면 shape 못쓰기도 하고 다음에 꺼내쓸때 오류날수있음
       'cuclaim_cn' : np.array(cuclaim_column_names),
       'tr_cucntt_dist' : np.array(train_cucntt_distri),
       'tr_cuclaim_dist' : np.array(train_cuclaim_distri),
      }
pickle.dump(save,f,pickle.HIGHEST_PROTOCOL)
f.close()

#check pickle
print("\npicklize finished")
with open(pickle_name,'r') as f:
    trainwb=pickle.load(f)
    #print(pickle_name)
    #for k,v in trainwb.items():
        #print(k,v.shape)


def pickleread(pickle_name):
    with open(pickle_name,'rb') as f:
        data=pickle.load(f)
        return data

def sheetmake(data,excel_name):
    book = openpyxl.Workbook()
    for dictitle , dictdata in data.items():
        dictdata=np.matrix(dictdata) #1차원 배열 있으면 shape 차원 하나라 오류나서.
        sheet=book.create_sheet(title=dictitle)
        for n_col in range(0,dictdata.shape[1]):
            for n_row in range(0,dictdata.shape[0]):
                input_value=np.asscalar(dictdata[n_row,n_col])          #python native 로 바꿔주는 코드. 이것과 아래줄 둘중하나 필수. 
                #input_value=dictdata[n_row,n_col]                           #str 오류날때 asscalar 빼면 될때있음
                sheet.cell(row=n_row+1,column=n_col+1).value=input_value    #엑셀에선 행,열 첫번호가 1 
            #sheet.column_dimensions[openpyxl.cell.get_column_letter(n_col+1)].width = 2.76 #컬럼 넓이 조절. 필요 없으면 빼기
        #print('making sheet : ',dictitle)
    sheet = book.get_sheet_by_name('Sheet') #select sheet named Sheet
    book.remove_sheet(sheet) #delete that sheet
    #print('saving data to excel...')
    book.save(excel_name)
    print ('finished, file saved : ',excel_name)


#엑셀이름  :        train) 피클이름 노드수 러닝레이트 트레인횟수 
# train) rand1 node100 LR0.5 try1 .xlsx

excel_name=resultname+'.xlsx' 
sheetmake(pickleread(pickle_name),excel_name)


#한번 시행당 파일 3개가 만들어진다. 하나는 텐서플로 variable 방식 w,b저장 확장자없는것,  pickle로 w,b 저장, 엑셀로 w,b 출력한 3가지다. 