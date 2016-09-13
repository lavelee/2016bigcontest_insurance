# -*- coding: utf-8 -*-
import pickle
import openpyxl
import numpy
import os

#피클의 모든 item을 가져옵니다. 태그 하나당 시트 하나가 만들어집니다.
#trained_w12b12_cucntt.pickle
#trained_w12b12_cuclaim.pickle

#pickle_name = 'trained_w12b12_cuclaim.pickle'
pickle_name = 'insurance_nullfix_adv_skima.pickle'

#엑셀이름


def pickleread(pickle_name):
    with open(pickle_name,'rb') as f:
        data=pickle.load(f)
        return data

def sheetmake(data,excel_name):
    book = openpyxl.Workbook()
    for dictitle , dictdata in data.items():
        dictdata=numpy.matrix(dictdata) #1차원 배열 있으면 shape 차원 하나라 오류나서.
        sheet=book.create_sheet(title=dictitle)
        for n_col in range(0,dictdata.shape[1]):
            for n_row in range(0,dictdata.shape[0]):
                #input_value=numpy.asscalar(dictdata[n_row,n_col])          #python native 로 바꿔주는 코드. 이것과 아래줄 둘중하나 필수. 
                input_value=dictdata[n_row,n_col]                           #str 오류날때 asscalar 빼면 될때있음
                sheet.cell(row=n_row+1,column=n_col+1).value=input_value    #엑셀에선 행,열 첫번호가 1 
            #sheet.column_dimensions[openpyxl.cell.get_column_letter(n_col+1)].width = 2.76 #컬럼 넓이 조절. 필요 없으면 빼기
        print 'making sheet : ',dictitle
    sheet = book.get_sheet_by_name('Sheet') #select sheet named Sheet
    book.remove_sheet(sheet) #delete that sheet
    print 'saving data to excel...'
    book.save(excel_name)
    print 'finished, file saved : ',excel_name


excel_name=pickle_name[:pickle_name.rfind(".pickle")]+'.xlsx' #rfind 해야 경로명에 pickle 말고 확장자 pickle 만 찾아냄

if os.path.isfile(excel_name): #이미 파일이 있으면 삭제함 #엑셀파일이 열려있으면 삭제도 못하고 오류남. 
	os.remove(excel_name)
	print('target excel file exists. continue after deleting')
sheetmake(pickleread(pickle_name),excel_name)


