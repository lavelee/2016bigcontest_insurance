# -*- coding: utf-8 -*-
#graphviz 라는 외부프로그램 사용. http://www.graphviz.org/Download_windows.php 에서 윈도우용 다운. 35메가쯤됨.
#그 후에 pydotplus pip로  설치 (순서 지켜야 dot.exe 가 생성됨. )
#그후에 또 dot.exe 존재하는 경로를 찾아서  윈도우 환경변수 path에 등록.
#path 등록후엔 재부팅 필수. 재부팅전엔 추가 안된걸로 나옴.
#너무 커서 pdf 가로길이가 일정수준 넘으면 뷰어가 포기해서 희게나옴. 그럴땐 max_depth제한 지정+컬럼명 길경우 rotate=true
from sklearn.ensemble import AdaBoostClassifier as abc
from sklearn.ensemble import RandomForestClassifier as rf
from sklearn import tree
from sklearn.externals.six import StringIO
import pydotplus
import pickle
import numpy as np
import sys
import openpyxl

picklename='clcntt_randfix01.pickle'
trainer_select = 'adaboost'
# trainer_select = 'tree'
# trainer_select = 'randomforest'
n_estimators=10
learning_rate=0.8
pdfname='test.pdf'
ifpdf=0

try : #변수를 외부에서 받아오는거 설정하고 없으면 pass
    picklename= sys.argv[1]
    trainer_select = sys.argv[2] #adaboost, tree, randomforest 3중 하나를 받아옴. 없으면 지정값으로
    if trainer_select =='adaboost':
        n_estimators=int(sys.argv[3])
        learning_rate=float(sys.argv[4])
        test_repeat=int(sys.argv[5])
        #현재 submit 본의 결과테스트는 adaboost 만 결과출력하도록.
        excel_name = picklename+' '+trainer_select+' est'+str(n_estimators)+' LR'+str(learning_rate)+' TR'+str(test_repeat)+' submit.xlsx'
        # print excel_name
    elif trainer_select =='randomforest':
        n_estimators=int(sys.argv[3])
    elif trainer_select =='tree':
        filename      =sys.argv[3]
    else:
        raise NameError('Wrong argv : only adaboost/tree/randomforest available')
except IndexError :
    pass

def f1Score(predictions, labels):
    #print 'predictions_raw=',predictions
    predictions=np.array(predictions)
    predictions=np.arange(2) == predictions[:,None].astype(np.float32)
    labels     =np.array(labels)
    labels     =np.arange(2) == labels[:,None].astype(np.float32)
    predictions = np.array(predictions==predictions.max(axis=1)[:,None],dtype='float32')
    predict = np.sum(predictions,0)
    real = np.sum(labels,0)
    correct_predict = np.sum(labels*(np.argmax(predictions, 1) == np.argmax(labels, 1))[:,None],0)
    precision = float(correct_predict[1])/predict[1]
    recall = float(correct_predict[1])/real[1]
    f1_score=2*precision*recall/(precision+recall)
    #print 'predictions =',predictions
    print'sin predict :',predict[1]
    print'sin correct predict :',correct_predict[1]
    print'sin real :',real[1]
    print'sin F1 score :',f1_score
    print 'result printing finished'
    return f1_score

def sheetmake(data,excel_name):
    book = openpyxl.Workbook()
    for dictitle , dictdata in data.items():
        dictdata=np.matrix(dictdata) #1차원 배열 있으면 shape 차원 하나라 오류나서.
        sheet=book.create_sheet(title=dictitle)
        for n_col in range(0,dictdata.shape[1]):
            for n_row in range(0,dictdata.shape[0]):
                input_value=np.asscalar(dictdata[n_row,n_col])          #python native 로 바꿔주는 코드. 이것과 아래줄 둘중하나 필수. 
                #input_value=dictdata[n_row,n_col]                           #str 오류날때 asscalar 빼면 될때있음
                sheet.cell(row=n_row+1,column=n_col+1).value=input_value    #엑셀에선 행,열 첫번호가 1 
            # sheet.column_dimensions[openpyxl.cell.get_column_letter(n_col+1)].width = 2.76 #컬럼 넓이 조절. 필요 없으면 빼기
        print('making sheet : ',dictitle)
    sheet = book.get_sheet_by_name('Sheet') #select sheet named Sheet
    book.remove_sheet(sheet) #delete that sheet
    print('saving data to excel...')
    book.save(excel_name)
    print 'finished, file saved : ',excel_name

with open(picklename,'rb') as f:
    data=pickle.load(f)
test_labels =data['test_label']
test_dataset =data['test_data']
train_labels =data['train_label']
train_dataset =data['train_data']
column_names =data['col_names']
submit_dataset =data['submit_data']
submit_custid =data['submit_custid']
del data

#조건별로 트레이너와 변수 세팅
if trainer_select =='adaboost':
    trainer = abc(n_estimators=n_estimators,learning_rate=learning_rate).fit(train_dataset,train_labels)
elif trainer_select =='randomforest':
    trainer = rf(n_estimators=n_estimators).fit(train_dataset,train_labels)
elif trainer_select =='tree': #tree 는 pdf 파일도 작성
    trainer = tree.DecisionTreeClassifier().fit(train_dataset,train_labels)
    if ifpdf==1: #필요하다면 pdf로 출력
        dot_data=StringIO()
        tree.export_graphviz(trainer,
                             out_file=dot_data,
                             feature_names=column_names,
                             class_names=['innocent','sin'],
                             filled=True, rounded=True,
                             impurity=False,max_depth=6,rotate=True
                             )
        graph = pydotplus.graph_from_dot_data(dot_data.getvalue())
        graph.write_pdf(pdfname)
    else:
        pass
else:   
    raise NameError('Wrong trainer select : only adaboost/tree/randomforest available')

#prediction 수행후 f1 출력
tr_prediction = trainer.predict(test_dataset)
f1Score(tr_prediction,test_labels)

#매 submit 결과 예측
submit_estimate = trainer.predict(submit_dataset)[:,None]
print(submit_custid.shape)
print(submit_estimate.shape)
submit_result = np.concatenate((submit_custid,submit_estimate),1)
# print submit_result
print 'submit set sin total : ',np.sum(submit_result,0)[1]

# 예측값을 엑셀로 저장
submit_result={'submit':submit_result} #엑셀은 dict 의 id 로 탭만드는등 동작하니 바꿔줌.
sheetmake(submit_result, excel_name) 
#todo   

print 'finished'
