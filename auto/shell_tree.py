# -*- coding: utf-8 -*-
#tensorflow 는 리눅스 필요해서 따로 쉘 만들고, tree 시리즈는 윈도우에서도 되니 그것만 처리하는 쉘파일. 

import subprocess
import os
import pickle
import MySQLdb
import openpyxl
import time
import numpy
t1=time.localtime() #시작시간
print 'started at', time.strftime('%y%m%d %Hh%Mm',t1)

#pickle file location 
pickle_folder = 'D:/sql_to_pickle_forshell/pickles/cuclcntt/'  #나중에 파일명 추가하기 용이하게 마지막에 / 를 추가해줘야한다. 
subpy = 'D:/sql_to_pickle_forshell/tree_trainers_forshell.py'    #실행할 py 파일. 한글경로 들어가면 안됨

#변화시킬 변수지정 : 
trainer_select=['tree','randomforest','adaboost']
n_estimators=[10,100,500]
learning_rate=[0.2, 0.5, 0.9]
test_repeat = 4

#폴더내 pickle 파일들 리스트 얻기
files_name = os.listdir(pickle_folder)
print 'pickle list : ',files_name,'\n' #파일명만 리스트 체크
files_wpath = [pickle_folder+item for item in files_name]  #os.path.join 쓰면 경로와 파일사이에 \\ 를 넣어주는데 난 / 를 쓰고싶어서 직접. 
#print files_wpath #전체경로 포함한 리스트 체크

def pcrfCollector(book):
#sin predict, sin correct predict, sin real, sin f1 score 부분을 텍스트에서 찾고 오른쪽의 값을 리턴한다
    targets =['\nsin predict : ','\nsin correct predict : ','\nsin real : ','\nsin F1 score : ','\nfinished']
    result =[] #위 targets 배열의 마지막은 stopper 로 실제 탐색에 사용되진 않지만 멈추는 위치 지정이다. 
    for i in range(0,len(targets)-1) : #마지막 배열은 ending point 라서 제거 
        start = book.find(targets[i])+len(targets[i])
        end = book.find(targets[i+1])
        #print(book[start:end])
        result.append(float(book[start:end]))
    #print result
    return result

# def where(h,i,j,k,imax=len(node_var),jmax=len(Lrate_var),kmax=n_test_per_var): #작성 중..
# #현재위치 셈. hmax는 필요가 없다
#     position = (h*imax*jmax*kmax +
#                     i*        jmax*kmax +
#                     j*                kmax +
#                     k+1                       ) #마지막 digit에는 +1 해줘야함.
#     return position


def pickleread(pickle_name):
    with open(pickle_name,'rb') as f:
        data=pickle.load(f)
        return data

def sheetmake(data,excel_name):
    book = openpyxl.Workbook()
    data=numpy.matrix(data) #1차원 배열 있으면 shape 차원 하나라 오류나서.
    sheet=book.create_sheet(title='result')
    for n_col in range(0,data.shape[1]):
        for n_row in range(0,data.shape[0]):
            #input_value=numpy.asscalar(data[n_row,n_col])          #python native 로 바꿔주는 코드. 이것과 아래줄 둘중하나 필수. 
            input_value=data[n_row,n_col]                           #str 오류날때 asscalar 빼면 될때있음
            sheet.cell(row=n_row+1,column=n_col+1).value=input_value    #엑셀에선 행,열 첫번호가 1 
        #sheet.column_dimensions[openpyxl.cell.get_column_letter(n_col+1)].width = 2.76 #컬럼 넓이 조절. 필요 없으면 빼기
    sheet = book.get_sheet_by_name('Sheet') #select sheet named Sheet
    book.remove_sheet(sheet) #delete that sheet
    print 'saving data to excel...'
    book.save(excel_name)
    print 'finished, file saved : ',excel_name

tree_total_test= len(files_name)*test_repeat
rand_total_test= len(files_name)* len(n_estimators)*test_repeat
ada_total_test = len(files_name)*len(n_estimators)*len(learning_rate)*test_repeat
total=tree_total_test+rand_total_test+ada_total_test
where=0

test_result = [['picklename','trainer','estimators','L rate','test repeat','predict','correct predict','real','F1 score']] #자료 넣을 컬럼, 9컬럼.

    # subpy 변수 받는 부분
    # picklename= sys.argv[1]
    # trainer_select = sys.argv[2] #adaboost, tree, randomforest 3중 하나를 받아옴. 없으면 지정값으로
    # if trainer_select =='adaboost':
    #     n_estimators=int(sys.argv[3])
    #     learning_rate=float(sys.argv[4])
    # elif trainer_select =='randomforest':
    #     n_estimators=int(sys.argv[3])
    # elif trainer_select `=='tree':
    #     filename      =sys.argv[3]

try:
    global where
    for i,ii in enumerate(files_name):
        data=pickleread(pickle_folder+ii)
        for j ,jj in enumerate(trainer_select): #트레이너 선택
            for m in range(0,test_repeat): #반복횟수
                if jj=='tree':
                    syscommand = 'python '+subpy+' "'+pickle_folder+ii+'" '+jj+' '+pickle_folder+ii[:ii.find('.pickle')]+' '+str(m)+'.pdf'
                    get=subprocess.check_output(syscommand, shell=True)
                    output = pcrfCollector(get)
                    output.insert(0,m)#한 조건의 몇번째 try 인지
                    output.insert(0,'-') #learning rate
                    output.insert(0,'-') #estimators
                    output.insert(0,jj) #트레이너 종류
                    output.insert(0,ii) #filename
                    test_result.append(output) #이것도 재할당 안해도 자동 적용되는거. 
                    where=where+1
                    print where,'/',total,'result :',output #현재 위치와 최대값 표시

                elif jj=='randomforest':
                    for k, kk in enumerate(n_estimators): 
                        syscommand = 'python '+subpy+' "'+pickle_folder+ii+'" '+jj+' '+str(kk)
                        get=subprocess.check_output(syscommand, shell=True)
                        output = pcrfCollector(get)
                        output.insert(0,m)#한 조건의 몇번째 try 인지
                        output.insert(0,'-') #learning rate
                        output.insert(0,kk) #estimators
                        output.insert(0,jj) #트레이너 종류
                        output.insert(0,ii) #filename
                        test_result.append(output) #이것도 재할당 안해도 자동 적용되는거.
                        where=where+1 
                        print where,'/',total,'result :',output #현재 위치와 최대값 표시

                elif jj=='adaboost':
                    for k, kk in enumerate(n_estimators):
                        for l, ll in enumerate(learning_rate):
                            syscommand = 'python '+subpy+' "'+pickle_folder+ii+'" '+jj+' '+str(kk)+' '+str(ll)
                            get=subprocess.check_output(syscommand, shell=True)
                            output = pcrfCollector(get)
                            output.insert(0,m)#한 조건의 몇번째 try 인지
                            output.insert(0,ll) #learning rate
                            output.insert(0,kk) #estimators
                            output.insert(0,jj) #트레이너 종류
                            output.insert(0,ii) #filename
                            test_result.append(output) #이것도 재할당 안해도 자동 적용되는거. 
                            where=where+1
                            print where,'/',total,'result :',output 

    sheetmake(test_result,pickle_folder+'folder-result.xlsx')

finally:
    t2=time.localtime() #끝난 시간
    print 'started at', time.strftime('%y%m%d %Hh%Mm',t1)
    print 'ends at',  time.strftime('%y%m%d %Hh%Mm',t2)


